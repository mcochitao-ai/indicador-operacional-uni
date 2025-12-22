import openpyxl

def processar_backlog_fluxo(filepath, dia):
    """
    Processa o arquivo Excel de backlog fluxo e extrai os valores
    Lê da aba "Dinâmicas", linhas 108-119, e pega o valor do dia ANTERIOR
    
    Args:
        filepath: Caminho do arquivo Excel de backlog fluxo
        dia: Número do dia atual (1-31) - será usado dia-1 para buscar o backlog
    
    Returns:
        dict: Dicionário com nome do CD como chave e valor do backlog fluxo
    """
    try:
        # Abrir o arquivo Excel (suprimindo warnings)
        import warnings
        warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
        
        wb = openpyxl.load_workbook(filepath, data_only=True, keep_vba=True)
        
        # Selecionar a aba "Dinâmicas"
        if 'Dinâmicas' not in wb.sheetnames:
            raise ValueError('Aba "Dinâmicas" não encontrada no arquivo de backlog fluxo')
        
        ws = wb['Dinâmicas']
        
        # Dicionário para armazenar backlog fluxo por CD
        backlog_fluxo = {}
        
        # Calcular o dia anterior
        dia_anterior = dia - 1
        if dia_anterior < 1:
            # Se for dia 1, não há dia anterior, retornar vazio
            wb.close()
            return {}
        
        # Encontrar a coluna do dia anterior
        # Linha 108 tem os cabeçalhos com as datas
        coluna_dia = None
        for col_idx in range(1, 50):  # Verificar até coluna 50
            cell = ws.cell(row=108, column=col_idx)
            if cell.value:
                # Verificar se é uma data ou número que corresponde ao dia anterior
                valor_celula = str(cell.value)
                # Pode ser formato de data ou apenas o número do dia
                if str(dia_anterior) in valor_celula or valor_celula == str(dia_anterior):
                    coluna_dia = col_idx
                    break
                # Verificar se é uma data completa (DD/MM/YYYY ou datetime)
                try:
                    from datetime import datetime
                    if isinstance(cell.value, datetime):
                        if cell.value.day == dia_anterior:
                            coluna_dia = col_idx
                            break
                    elif '/' in valor_celula:
                        # Formato DD/MM/YYYY
                        dia_na_celula = int(valor_celula.split('/')[0])
                        if dia_na_celula == dia_anterior:
                            coluna_dia = col_idx
                            break
                except:
                    pass
        
        if coluna_dia is None:
            # Se não encontrou o dia anterior, retornar vazio
            wb.close()
            return {}
        
        # Ler os dados das linhas 109 a 119 (CDs estão aqui)
        for row in range(109, 120):  # 109 a 119
            # Coluna A ou B deve ter o nome do CD (vou tentar ambas)
            cd_nome = ws.cell(row=row, column=1).value  # Coluna A
            if not cd_nome or cd_nome == '' or cd_nome == 'Total Geral':
                cd_nome = ws.cell(row=row, column=2).value  # Tentar coluna B
            
            # Pular linhas vazias ou Total Geral
            if not cd_nome or cd_nome == '' or cd_nome == 'Total Geral':
                continue
            
            # Valor do backlog fluxo na coluna do dia anterior
            valor_backlog = ws.cell(row=row, column=coluna_dia).value
            
            valor_formatado = 0
            if valor_backlog is not None:
                try:
                    valor_formatado = float(valor_backlog)
                except (ValueError, TypeError):
                    valor_formatado = 0
            
            # Normalizar nome do CD para garantir correspondência
            cd_nome_normalizado = str(cd_nome).strip().upper()
            
            # Mapear nomes possíveis
            mapeamento_nomes = {
                'CABO STO AGOSTINHO': 'CABO DE SANTO AGOSTINHO',
                'CABO': 'CABO DE SANTO AGOSTINHO',
                'GOIÂNIA': 'GOIÂNIA',
                'IGARASSU': 'IGARASSU',
                'INDAIATUBA': 'INDAIATUBA',
                'JABOATÃO': 'JABOATÃO',
                'LOUVEIRA': 'LOUVEIRA',
                'POUSO ALEGRE': 'POUSO ALEGRE',
                'SERRA': 'SERRA'
            }
            
            # Usar o nome mapeado se existir, senão usar o nome original
            cd_final = mapeamento_nomes.get(cd_nome_normalizado, cd_nome)
            
            backlog_fluxo[cd_final] = valor_formatado
        
        wb.close()
        
        return backlog_fluxo
        
    except Exception as e:
        raise Exception(f'Erro ao processar arquivo de backlog fluxo: {str(e)}')


def processar_capacidade(filepath, dia):
    """
    Processa o arquivo Excel e extrai dados de capacidade dos CDs
    
    Args:
        filepath: Caminho do arquivo Excel
        dia: Número do dia (1-31) correspondente à aba
    
    Returns:
        dict: Dicionário com dados dos CDs
    """
    try:
        # Abrir o arquivo Excel (incluindo .xlsm com macros)
        wb = openpyxl.load_workbook(filepath, data_only=True, keep_vba=True)
        
        # Selecionar a aba do dia
        nome_aba = str(dia)
        if nome_aba not in wb.sheetnames:
            raise ValueError(f'Aba "{nome_aba}" não encontrada no arquivo')
        
        ws = wb[nome_aba]
        
        # Extrair dados dos CDs (linhas 4 a 11, coluna B)
        cds = []
        
        for row in range(4, 12):  # 4 a 11 (incluindo 11)
            cd_nome = ws[f'B{row}'].value
            
            # Pular linhas vazias
            if not cd_nome:
                continue
            
            # Capacidade geral: X / C
            valor_x = ws[f'X{row}'].value
            valor_c = ws[f'C{row}'].value
            
            # Capacidade de pallet: AM
            capacidade_pallet = ws[f'AM{row}'].value
            
            # Capacidade de caixas: AH
            capacidade_caixas = ws[f'AH{row}'].value
            
            # Status de abertura para inclusão: Y
            status_inclusao = ws[f'Y{row}'].value
            
            # Status de abertura para caixas: Z
            status_caixas = ws[f'Z{row}'].value
            
            # Status de abertura para pallets: AA
            status_pallets = ws[f'AA{row}'].value
            
            # Dock Vendas para Faturamento: D
            dock_vendas = ws[f'D{row}'].value
            
            # Inclusão: E
            inclusao = ws[f'E{row}'].value
            
            # Dock Total - Vendas: F e I
            vendas_f = ws[f'F{row}'].value
            vendas_i = ws[f'I{row}'].value
            
            # Dock Total - Transferências: N, Q, T, U
            transf_n = ws[f'N{row}'].value
            transf_q = ws[f'Q{row}'].value
            transf_t = ws[f'T{row}'].value
            transf_u = ws[f'U{row}'].value
            
            # Agendamentos: AB
            agendamentos = ws[f'AB{row}'].value
            
            # Backlog de Transferências: S (linhas 4-11)
            backlog_transferencias = ws[f'S{row}'].value
            
            # Calcular capacidade geral (sempre será X/C * 100)
            capacidade_geral = None
            if valor_x is not None and valor_c is not None and valor_c != 0:
                try:
                    capacidade_geral = round((float(valor_x) / float(valor_c)) * 100, 0)
                except (ValueError, TypeError):
                    capacidade_geral = None
            
            # Formatar valores de pallet e caixas
            # Assumir que valores <= 10 são decimais (ex: 0.865 ou 1.06) e precisam ser multiplicados por 100
            # Valores > 10 já estão em percentual
            cap_pallet_formatada = None
            cap_caixas_formatada = None
            
            if capacidade_pallet is not None:
                try:
                    valor_float = float(capacidade_pallet)
                    # Se valor <= 10, provavelmente é decimal (0.865 = 86.5% ou 1.06 = 106%)
                    if valor_float <= 10:
                        cap_pallet_formatada = round(valor_float * 100, 0)
                    else:
                        cap_pallet_formatada = round(valor_float, 0)
                except (ValueError, TypeError):
                    cap_pallet_formatada = None
            
            if capacidade_caixas is not None:
                try:
                    valor_float = float(capacidade_caixas)
                    # Se valor <= 10, provavelmente é decimal (0.011 = 1.1% ou 1.06 = 106%)
                    if valor_float <= 10:
                        cap_caixas_formatada = round(valor_float * 100, 0)
                    else:
                        cap_caixas_formatada = round(valor_float, 0)
                except (ValueError, TypeError):
                    cap_caixas_formatada = None
            
            # Formatar dock vendas e inclusão
            dock_vendas_formatado = None
            inclusao_formatada = None
            total_com_inclusao = None
            
            if dock_vendas is not None:
                try:
                    dock_vendas_formatado = float(dock_vendas)
                except (ValueError, TypeError):
                    dock_vendas_formatado = 0
            else:
                dock_vendas_formatado = 0
            
            if inclusao is not None and inclusao != "" and inclusao != 0:
                try:
                    inclusao_formatada = float(inclusao)
                except (ValueError, TypeError):
                    inclusao_formatada = None
            
            # Calcular total com inclusão
            if dock_vendas_formatado is not None:
                if inclusao_formatada is not None and inclusao_formatada != 0:
                    total_com_inclusao = dock_vendas_formatado + inclusao_formatada
                else:
                    total_com_inclusao = dock_vendas_formatado
            
            # Processar Dock Total - Vendas
            dock_total_vendas = 0
            if vendas_f is not None:
                try:
                    dock_total_vendas += float(vendas_f)
                except (ValueError, TypeError):
                    pass
            if vendas_i is not None:
                try:
                    dock_total_vendas += float(vendas_i)
                except (ValueError, TypeError):
                    pass
            
            # Processar Dock Total - Transferências
            dock_total_transferencias = 0
            for valor in [transf_n, transf_q, transf_t, transf_u]:
                if valor is not None:
                    try:
                        dock_total_transferencias += float(valor)
                    except (ValueError, TypeError):
                        pass
            
            # Total Dock
            dock_total_geral = dock_total_vendas + dock_total_transferencias
            
            # Total de Fluxo (coluna G)
            total_fluxo = ws[f'G{row}'].value
            total_fluxo_formatado = 0
            if total_fluxo is not None:
                try:
                    total_fluxo_formatado = float(total_fluxo)
                except (ValueError, TypeError):
                    total_fluxo_formatado = 0
            
            # Agendamentos formatado
            agendamentos_formatado = 0
            if agendamentos is not None:
                try:
                    agendamentos_formatado = float(agendamentos)
                except (ValueError, TypeError):
                    agendamentos_formatado = 0
            
            # Backlog de Transferências formatado
            backlog_transferencias_formatado = 0
            if backlog_transferencias is not None:
                try:
                    backlog_transferencias_formatado = float(backlog_transferencias)
                except (ValueError, TypeError):
                    backlog_transferencias_formatado = 0
            
            cd_data = {
                'nome': cd_nome,
                'capacidade_geral': capacidade_geral,
                'capacidade_pallet': cap_pallet_formatada,
                'capacidade_caixas': cap_caixas_formatada,
                'status_inclusao': status_inclusao,
                'status_caixas': status_caixas,
                'status_pallets': status_pallets,
                'dock_vendas': dock_vendas_formatado,
                'inclusao': inclusao_formatada,
                'total_com_inclusao': total_com_inclusao,
                'dock_total_vendas': dock_total_vendas,
                'dock_total_transferencias': dock_total_transferencias,
                'dock_total_geral': dock_total_geral,
                'agendamentos': agendamentos_formatado,
                'backlog_transferencias': backlog_transferencias_formatado,
                'total_fluxo': total_fluxo_formatado
            }
            
            cds.append(cd_data)
        
        # Extrair Perdas W e T (linhas 16 a 23, colunas G e H)
        # E adicionar backlog_vendas, backlog_expedido e backlog_total aos CDs
        perdas = []
        for idx, cd in enumerate(cds):
            row = 16 + idx  # linha 16 corresponde ao primeiro CD, 17 ao segundo, etc.
            
            perda_w = ws[f'G{row}'].value
            perda_t = ws[f'H{row}'].value
            faturado_expedido = ws[f'I{row}'].value
            
            perda_w_formatada = 0
            perda_t_formatada = 0
            faturado_expedido_formatado = 0
            
            if perda_w is not None:
                try:
                    perda_w_formatada = float(perda_w)
                except (ValueError, TypeError):
                    perda_w_formatada = 0
            
            if perda_t is not None:
                try:
                    perda_t_formatada = float(perda_t)
                except (ValueError, TypeError):
                    perda_t_formatada = 0
            
            if faturado_expedido is not None:
                try:
                    faturado_expedido_formatado = float(faturado_expedido)
                except (ValueError, TypeError):
                    faturado_expedido_formatado = 0
            
            # Backlog de vendas = (Perdas W + Perdas T) - Faturado/Expedido
            backlog_vendas = (perda_w_formatada + perda_t_formatada) - faturado_expedido_formatado
            # Garantir que não seja negativo
            if backlog_vendas < 0:
                backlog_vendas = 0
            
            # Fluxo Fiscal = Total de Fluxo - Faturado/Expedido
            fluxo_fiscal = cd['total_fluxo'] - faturado_expedido_formatado
            # Garantir que não seja negativo
            if fluxo_fiscal < 0:
                fluxo_fiscal = 0
            
            # Backlog total = Backlog de vendas + Backlog de transferências
            backlog_total = backlog_vendas + cd['backlog_transferencias']
            
            # Adicionar backlog_vendas, backlog_expedido, fluxo_fiscal e backlog_total ao CD
            cd['backlog_vendas'] = backlog_vendas
            cd['backlog_expedido'] = faturado_expedido_formatado
            cd['fluxo_fiscal'] = fluxo_fiscal
            cd['backlog_total'] = backlog_total
            
            perda_data = {
                'nome': cd['nome'],
                'perda_w': perda_w_formatada,
                'perda_t': perda_t_formatada,
                'backlog': backlog_vendas  # Manter compatibilidade
            }
            
            perdas.append(perda_data)
        
        wb.close()
        
        return {
            'dia': dia,
            'cds': cds,
            'perdas': perdas
        }
        
    except Exception as e:
        raise Exception(f'Erro ao processar arquivo: {str(e)}')
